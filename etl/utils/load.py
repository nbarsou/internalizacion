import json
import logging
import hashlib
import unicodedata
import re
import uuid
import polars as pl
import openpyxl
from pathlib import Path
from typing import Any, Tuple

log = logging.getLogger(__name__)


# ============================================================================
# JSON
# ============================================================================


def load_json(path: Path) -> Any:
    """
    Load a JSON file. Raises on missing file or bad JSON so the
    caller can decide how to handle it — never silently returns {}.
    """
    log.debug(f"Loading JSON: {path}")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    log.debug(f"  → {len(data)} items")
    return data


# ============================================================================
# HELPERS
# ============================================================================


def _get_file_hash(path: Path) -> str:
    """SHA-1 hash of a file — used to detect Excel changes between runs."""
    h = hashlib.sha1()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            h.update(chunk)
    return h.hexdigest()


def _clean_header(name: str) -> str:
    """
    'ÁREA, ESCUELA O FACULTAD' → 'area_escuela_o_facultad'
    Strips accents, lowercases, collapses spaces to underscores,
    removes every character that is not [a-z0-9_].
    """
    if not name:
        return f"col_{uuid.uuid4().hex[:8]}"
    nfkd = unicodedata.normalize("NFKD", str(name))
    no_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
    clean = no_accents.lower().strip().replace(" ", "_")
    return re.sub(r"[^a-z0-9_]", "", clean)


def _clean_excel_comment(text: str) -> str:
    """
    Strip the Microsoft Threaded Comment boilerplate that openpyxl surfaces
    when a cell has a modern (threaded) comment.

    Two variants exist in this Excel:
      - Spanish: [Comentario encadenado] ... linkid=870924 ... Comentario:
      - English: [Threaded Comment] ... linkid=870924 ... Comment:

    Both share the linkid=870924 anchor so one regex handles both.
    Falls back to title-casing just the Comment: payload so type names are
    consistently capitalised regardless of how the author typed them.
    """
    if not text:
        return ""
    text = str(text).strip()
    if "linkid=870924" in text.lower():
        # Match either language variant, capture only the payload after "Comment:"
        pattern = r"(?si)\[.*?(?:encadenado|threaded comment)\].*?linkid=870924.*?(?:Comentario|Comment):\s*"
        cleaned = re.sub(pattern, "", text).strip()
        return cleaned
    return text


# ============================================================================
# CORE EXTRACTION
# ============================================================================


def _extract_excel(excel_path: Path) -> Tuple[pl.DataFrame, pl.DataFrame]:
    """
    Read the raw Excel file and return two DataFrames:
      df_main     — one row per university, headers normalised to snake_case,
                    hyperlink targets extracted as separate *_url columns,
                    a stable UUID added as row_id.
      df_comments — one row per cell comment, linked back via row_id.

    Does NOT touch config or cache — pure extraction logic.
    """
    log.info(f"Reading Excel: {excel_path}")

    # ── Polars read (all strings — no type inference surprises) ──────────────
    df_raw = pl.read_excel(excel_path, infer_schema_length=0)
    rename_map = {col: _clean_header(col) for col in df_raw.columns}
    df_raw = df_raw.rename(rename_map)

    # Drop any phantom header row — happens when the sheet has a blank or
    # merged row above the real header, causing Polars to treat the real
    # header row as data. The giveaway: the universidad cell equals the
    # literal column name "universidad" after normalisation.
    before = df_raw.height
    df_raw = df_raw.filter(
        pl.col("universidad").str.to_lowercase().str.strip_chars().ne("universidad")
    )
    dropped = before - df_raw.height
    if dropped:
        log.warning(f"Dropped {dropped} phantom header row(s) from Excel data")

    row_ids = [str(uuid.uuid4()) for _ in range(df_raw.height)]
    df_raw = df_raw.with_columns(pl.Series("row_id", row_ids))

    log.debug(f"Columns after normalisation: {df_raw.columns}")

    # ── openpyxl pass — hyperlinks + comments ────────────────────────────────
    wb = openpyxl.load_workbook(excel_path, data_only=False, keep_vba=False)
    ws = wb.active

    header_map: dict[int, str] = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            header_map[idx] = _clean_header(cell.value)

    url_updates: dict[tuple[int, str], str] = {}
    comments_list: list[dict] = []

    for i, (excel_row, row_uuid) in enumerate(zip(ws.iter_rows(min_row=2), row_ids)):
        for col_idx, cell in enumerate(excel_row):
            if col_idx not in header_map:
                continue
            col_name = header_map[col_idx]

            if cell.hyperlink and cell.hyperlink.target:
                url_updates[(i, col_name)] = cell.hyperlink.target

            if cell.comment:
                comments_list.append(
                    {
                        "row_id": row_uuid,
                        "column_source": col_name,
                        "comment_text": _clean_excel_comment(cell.comment.text),
                        "author": cell.comment.author,
                    }
                )

    # ── Attach hyperlink columns ──────────────────────────────────────────────
    url_cols = set(col for (_, col) in url_updates)
    df_main = df_raw.with_columns(
        [
            pl.Series(
                f"{col}_url",
                [url_updates.get((r, col)) for r in range(df_raw.height)],
            )
            for col in url_cols
        ]
    )

    # ── Global whitespace / newline cleanup ───────────────────────────────────
    df_main = df_main.with_columns(
        pl.col(pl.String).str.replace_all(r"[\r\n]+", " ").str.strip_chars()
    )

    # ── Comments DataFrame ────────────────────────────────────────────────────
    df_comments = (
        pl.from_dicts(comments_list)
        if comments_list
        else pl.DataFrame(
            {"row_id": [], "column_source": [], "comment_text": [], "author": []}
        )
    )

    return df_main, df_comments


# ============================================================================
# PUBLIC API
# ============================================================================


def get_data(
    excel_path: Path,
    cache_dir: Path,
    force_refresh: bool = False,
) -> Tuple[pl.DataFrame, pl.DataFrame]:
    """
    Return (df_main, df_comments) — from Parquet cache when possible,
    re-extracted from Excel when the file has changed or force_refresh=True.

    Args:
        excel_path:    Absolute path to the source .xlsx file.
        cache_dir:     Directory where Parquet files and the hash file live.
        force_refresh: Skip cache check and always re-process.

    Raises:
        FileNotFoundError: If excel_path does not exist.
        ValueError:        If the resulting DataFrame has no rows.
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    cache_dir.mkdir(parents=True, exist_ok=True)

    parquet_main = cache_dir / "cache_universities_silver.parquet"
    parquet_comments = cache_dir / "cache_comments_silver.parquet"
    hash_file = cache_dir / ".excel_hash"

    current_hash = _get_file_hash(excel_path)
    stored_hash = hash_file.read_text().strip() if hash_file.exists() else ""
    cache_valid = (
        not force_refresh
        and parquet_main.exists()
        and parquet_comments.exists()
        and current_hash == stored_hash
    )

    if cache_valid:
        log.info("Cache hit — loading from Parquet")
        df_main = pl.read_parquet(parquet_main)
        df_comments = pl.read_parquet(parquet_comments)
    else:
        reason = "force_refresh=True" if force_refresh else "Excel changed or no cache"
        log.info(f"Cache miss ({reason}) — extracting from Excel")
        df_main, df_comments = _extract_excel(excel_path)
        df_main.write_parquet(parquet_main)
        df_comments.write_parquet(parquet_comments)
        hash_file.write_text(current_hash)
        log.info("Cache updated")

    # ── Validation ────────────────────────────────────────────────────────────
    if df_main.height == 0:
        raise ValueError("Excel loaded but contains no data rows — check the file.")

    log.info(f"Loaded {df_main.height} rows, {df_main.width} columns")
    log.debug(f"Columns: {df_main.columns}")

    if df_comments.height == 0:
        log.warning("No cell comments found — df_comments is empty")
    else:
        log.info(f"Loaded {df_comments.height} cell comments")

    return df_main, df_comments
