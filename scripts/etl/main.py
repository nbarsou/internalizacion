import pandas as pd
import unicodedata
from openpyxl import load_workbook
import re

convenios = "data/conveniosinternacionales.xlsx"
referencias = "data/referencias.xlsx"
archivo_salida = "universidad.xlsx"

columnasconvenios = [
    "Universidad", "INICIO", "VIGENCIA", "CIUDAD", "¿Católica?",
    "¿SE USA?", "Página Web", "Contacto Intercambios Entrantes",
    "Contacto Intercambios Salientes", "Link convenio ",
    "País", "Región", "GIRO", "CAMPUS TITULAR",
    "STUDY ABROAD", "INTERCAMBIO", "PLAZAS POSGRADO",
    "PRÁCTICAS", "INVESTIGACIÓN", "DOBLE DIPLOMA", "COTUTELA",
    "PLAZAS LIC", "OTRO",
    "ÁREA, ESCUELA O FACULTAD"
]

dfconvenio = pd.read_excel(convenios, nrows=424)
dfreferencia = pd.read_excel(referencias)

dfconvenio = dfconvenio[columnasconvenios]

wb = load_workbook(convenios)
ws = wb.active

# -----------------------------
# LIMPIAR UNIVERSIDADES
# -----------------------------

dfconvenio = dfconvenio[dfconvenio["Universidad"].notna()]
dfconvenio = dfconvenio[dfconvenio["Universidad"].astype(str).str.strip() != ""]

dfconvenio.insert(0, "id_universidad", range(1, len(dfconvenio) + 1))

# NUEVO: católica default NO
dfconvenio["¿Católica?"] = dfconvenio["¿Católica?"].fillna("NO")

# -----------------------------
# NORMALIZAR
# -----------------------------

def normalizar(texto):
    if not isinstance(texto, str):
        return ""
    texto = texto.lower().strip()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return texto

# -----------------------------
# NORMALIZAR REFERENCIAS
# -----------------------------

dfconvenio["pais_norm"] = dfconvenio["País"].map(normalizar)
dfconvenio["region_norm"] = dfconvenio["Región"].map(normalizar)
dfconvenio["giro_norm"] = dfconvenio["GIRO"].map(normalizar)
dfconvenio["campus_norm"] = dfconvenio["CAMPUS TITULAR"].map(normalizar)

dfreferencia["pais_norm"] = dfreferencia["País"].map(normalizar)
dfreferencia["region_norm"] = dfreferencia["Región"].map(normalizar)
dfreferencia["giro_norm"] = dfreferencia["Giro"].map(normalizar)
dfreferencia["campus_norm"] = dfreferencia["campus"].map(normalizar)

# -----------------------------
# CATÁLOGO TIPO
# -----------------------------

df_tipo = dfreferencia[["tipo", "tipo_id"]].dropna().drop_duplicates()
df_tipo["tipo_norm"] = df_tipo["tipo"].map(normalizar)

# 🔥 SOLO TIPOS DINÁMICOS (>=7)
df_tipo_dinamico = df_tipo[df_tipo["tipo_id"] >= 7].copy()

mapa_tipo = dict(zip(df_tipo_dinamico["tipo_norm"], df_tipo_dinamico["tipo_id"]))

def obtener_id_tipo_dinamico(texto):
    global df_tipo, df_tipo_dinamico, mapa_tipo

    texto_norm = normalizar(texto)

    # 🔥 SOLO COMPARA CONTRA TIPOS >= 7
    for _, row in df_tipo_dinamico.iterrows():
        tipo_norm = row["tipo_norm"]
        primera = tipo_norm.split()[0] if tipo_norm else ""

        if primera and primera in texto_norm:
            return row["tipo_id"]

    # 🔥 GENERAR NUEVO ID CORRECTO
    max_id = df_tipo["tipo_id"].max() if not df_tipo.empty else 6
    nuevo_id = max(max_id + 1, 7)

    nueva_fila = pd.DataFrame([{
        "tipo": texto,
        "tipo_id": nuevo_id,
        "tipo_norm": texto_norm
    }])

    df_tipo = pd.concat([df_tipo, nueva_fila], ignore_index=True)
    df_tipo_dinamico = pd.concat([df_tipo_dinamico, nueva_fila], ignore_index=True)

    mapa_tipo[texto_norm] = nuevo_id

    return nuevo_id

# -----------------------------
# CATÁLOGO ESCUELAS
# -----------------------------

df_escuela = dfreferencia[["CVE", "Descripción"]].dropna().drop_duplicates()
df_escuela.columns = ["cve", "escuela"]
df_escuela["escuela_norm"] = df_escuela["escuela"].map(normalizar)

cve_existentes = df_escuela["cve"].tolist()
cve_numericos = [int(c) for c in cve_existentes if str(c).isdigit()]
nuevo_cve = max(cve_numericos, default=0) + 1

# -----------------------------
# CATÁLOGOS REFERENCIAS
# -----------------------------

df_region = dfreferencia[["Región","region_norm","ID_region"]].dropna().drop_duplicates()
df_pais = dfreferencia[["País","pais_norm","ID_pais"]].dropna().drop_duplicates()
df_giro = dfreferencia[["Giro","giro_norm","ID_giro"]].dropna().drop_duplicates()
df_campus = dfreferencia[["campus","campus_norm","campus_id"]].dropna().drop_duplicates()

# -----------------------------
# FUNCIONES AUXILIARES
# -----------------------------

def cumple(valor):
    if pd.isna(valor):
        return False
    valor = str(valor).strip().lower()
    return not (valor == "" or valor.startswith("n"))

def separar_escuelas(texto):
    if not isinstance(texto, str):
        return []
    partes = re.split(r'[,/]', texto)
    return [p.strip() for p in partes if p.strip()]

def obtener_numero_posgrado(texto):
    if not texto:
        return ""
    texto = texto.lower()
    if "posgrado" in texto:
        antes = texto.split("posgrado")[0]
        nums = re.findall(r'\d+', antes)
        if nums:
            return int(nums[-1])
    return ""

def obtener_numero(texto, posicion=2):
    if texto:
        nums = re.findall(r'\d+', texto)
        if len(nums) >= posicion:
            return int(nums[posicion-1])
    return ""

def extraer_texto_otro(valor, comentario):
    if comentario:
        texto = comentario.lower()
        if ":" in texto:
            return texto.split(":")[-1].strip()
        return texto.strip()
    return str(valor)

def extraer_final_comentario(texto):
    if not texto:
        return ""
    texto = texto.strip()
    if ":" in texto:
        return texto.split(":")[-1].strip()
    return texto

# -----------------------------
# FUNCIÓN REFERENCIAS
# -----------------------------

def actualizar_referencias(df_conv, df_ref, col_original, col_norm, ref_original, ref_norm, col_id):

    mapa = dict(zip(df_ref[ref_norm], df_ref[col_id]))

    faltantes = (
        df_conv.loc[~df_conv[col_norm].isin(df_ref[ref_norm]), col_norm]
        .dropna()
        .unique()
    )

    if len(faltantes) > 0:

        nuevo_id = df_ref[col_id].dropna().astype(int).max() + 1 if df_ref[col_id].notna().any() else 1

        nuevas_filas = []

        for valor_norm in faltantes:

            valor_original = df_conv.loc[df_conv[col_norm] == valor_norm, col_original].iloc[0]

            nuevas_filas.append({
                ref_original: valor_original,
                ref_norm: valor_norm,
                col_id: nuevo_id
            })

            mapa[valor_norm] = nuevo_id
            nuevo_id += 1

        df_ref = pd.concat([df_ref, pd.DataFrame(nuevas_filas)], ignore_index=True)

    df_conv[col_id] = df_conv[col_norm].map(mapa)

    return df_conv, df_ref

# -----------------------------
# PROCESAMIENTO
# -----------------------------

filas_convenio = []
filas_convenio_cve = []
filas_observaciones = []
filas_contacto = []

id_convenio = 1

# -----------------------------
# MAPEAR REFERENCIAS
# -----------------------------

dfconvenio, df_pais = actualizar_referencias(
    dfconvenio, df_pais,
    "País","pais_norm",
    "País","pais_norm",
    "ID_pais"
)

dfconvenio, df_region = actualizar_referencias(
    dfconvenio, df_region,
    "Región","region_norm",
    "Región","region_norm",
    "ID_region"
)

dfconvenio, df_giro = actualizar_referencias(
    dfconvenio, df_giro,
    "GIRO","giro_norm",
    "Giro","giro_norm",
    "ID_giro"
)

dfconvenio, df_campus = actualizar_referencias(
    dfconvenio, df_campus,
    "CAMPUS TITULAR","campus_norm",
    "campus","campus_norm",
    "campus_id"
)

for idx, row in dfconvenio.iterrows():

    id_uni = row["id_universidad"]
    # -------- CONTACTOS --------

    contacto_entrante = row["Contacto Intercambios Entrantes"]
    contacto_saliente = row["Contacto Intercambios Salientes"]

    def limpiar_contacto(valor):
        if pd.isna(valor):
            return ""
        valor = str(valor).strip()
        return valor if valor != "" else ""

    contacto_entrante = limpiar_contacto(contacto_entrante)
    contacto_saliente = limpiar_contacto(contacto_saliente)

    # Caso: ambos existen
    if contacto_entrante and contacto_saliente:

        if contacto_entrante == contacto_saliente:
            # mismo contacto → ESA 3
            filas_contacto.append([id_uni, contacto_entrante, 3])
        else:
            # diferentes → ESA 1 y 2
            filas_contacto.append([id_uni, contacto_entrante, 1])
            filas_contacto.append([id_uni, contacto_saliente, 2])

    # Solo entrante
    elif contacto_entrante:
        filas_contacto.append([id_uni, contacto_entrante, 1])

    # Solo saliente
    elif contacto_saliente:
        filas_contacto.append([id_uni, contacto_saliente, 2])
        link = row["Link convenio "]
        excel_row = idx + 2

    link = row["Link convenio "]
    excel_row = idx + 2

    # -------- OBSERVACIONES --------
    comentarios = []

    for col in ["A", "C", "E", "N"]:
        cell = ws[f"{col}{excel_row}"]
        if cell.comment:
            texto = extraer_final_comentario(cell.comment.text)
            if texto:
                comentarios.append(texto)

    cell_L = ws[f"L{excel_row}"]
    if cell_L.comment:
        texto_L = cell_L.comment.text.lower()
        if "doctorado" not in texto_L and "posgrado" not in texto_L:
            texto = extraer_final_comentario(cell_L.comment.text)
            if texto:
                comentarios.append(texto)

    cell_S = ws[f"S{excel_row}"]
    if cell_S.comment:
        texto_S = cell_S.comment.text.lower()
        if "posgrado" in texto_S:
            texto = extraer_final_comentario(cell_S.comment.text)
            if texto:
                comentarios.append(texto)

    observaciones = ". ".join(comentarios)

    comentario_lic = ws[f"L{excel_row}"].comment
    comentario_lic = comentario_lic.text.lower() if comentario_lic else ""

    comentario_otro = ws[f"S{excel_row}"].comment
    comentario_otro = comentario_otro.text if comentario_otro else ""

    valor_otro = row["OTRO"]

    # -------- ESCUELAS --------
    escuelas_texto = separar_escuelas(row["ÁREA, ESCUELA O FACULTAD"])
    cves_encontrados = set()

    for esc in escuelas_texto:
        esc_norm = normalizar(esc)

        if "abierto" in esc_norm:
            cves_encontrados.add("ZZ")
            continue
        
        if "estudios" in esc_norm:
            cves_encontrados.add("EG")
            continue

        if "deporte" in esc_norm:
            cves_encontrados.add("CD")
            continue

        if "comunicacion" in esc_norm:
            cves_encontrados.add("CC")
            continue

        encontrado = False

        for _, ref in df_escuela.iterrows():
            if ref["escuela_norm"] in esc_norm:
                cves_encontrados.add(ref["cve"])
                encontrado = True

        if not encontrado and esc_norm != "":
            cves_encontrados.add(nuevo_cve)

            df_escuela = pd.concat([
                df_escuela,
                pd.DataFrame([{
                    "cve": nuevo_cve,
                    "escuela": esc,
                    "escuela_norm": esc_norm
                }])
            ], ignore_index=True)

            nuevo_cve += 1

    plazas_posgrado_extra = obtener_numero_posgrado(comentario_lic)

    nivel_intercambio = 1
    if "doctorado" in comentario_lic:
        nivel_intercambio = 3

    # -------- TIPOS --------

    def agregar_convenio(tipo_id, plazas, niveles):
        global id_convenio

        for n in niveles:

            filas_convenio.append([
                id_convenio,
                id_uni,
                tipo_id,
                plazas,
                link,
                n  # 🔥 NUEVA COLUMNA nivel
         ])

            for cve in cves_encontrados:
                filas_convenio_cve.append([id_convenio, cve])

            if observaciones:
                filas_observaciones.append([id_convenio, observaciones])

            id_convenio += 1

    if cumple(row["STUDY ABROAD"]):
        agregar_convenio(1, "", [1,2,3])

    if cumple(row["INTERCAMBIO"]):
        plazas = row["PLAZAS LIC"]
        plazas = int(plazas) if str(plazas).isdigit() else ""

        if plazas_posgrado_extra != "" and isinstance(plazas, int):
            plazas -= plazas_posgrado_extra

        agregar_convenio(2, plazas, [nivel_intercambio])

    if cumple(row["PLAZAS POSGRADO"]):
        if plazas_posgrado_extra != "":
            plazas = plazas_posgrado_extra
        else:
            plazas = row["PLAZAS POSGRADO"]
            plazas = int(plazas) if str(plazas).isdigit() else ""

        agregar_convenio(2, plazas, [2,3])

    if cumple(row["PRÁCTICAS"]):
        agregar_convenio(3, "", [1,2,3])

    if cumple(row["INVESTIGACIÓN"]):
        cell = ws[f"O{excel_row}"]
        comentario = cell.comment.text if cell.comment else ""
        plazas = obtener_numero(comentario, 2)

        agregar_convenio(4, plazas, [1,2,3])

    if cumple(row["DOBLE DIPLOMA"]):
        agregar_convenio(5, "", [1,2,3])

    if cumple(row["COTUTELA"]):
        agregar_convenio(6, "", [3])

    if cumple(valor_otro):

        valor_limpio = str(valor_otro).strip().lower()
        texto = extraer_texto_otro(valor_otro, comentario_otro)
        texto_norm = normalizar(texto)

        if "posgrado" not in texto_norm:

            if (valor_limpio in ["x", "mou"]) and not comentario_otro:
                tipo_id = 7
            else:
                tipo_id = obtener_id_tipo_dinamico(texto)

            agregar_convenio(tipo_id, "", [1,2,3])

# -----------------------------
# DATAFRAMES
# -----------------------------

df_convenio = pd.DataFrame(
    filas_convenio,
    columns=[
        "id_convenio",
        "id_universidad",
        "id_tipo",
        "plazas",
        "link_convenio",
        "nivel"
    ]
)

df_conv_cve = pd.DataFrame(
    filas_convenio_cve,
    columns=["id_convenio", "cve"]
)

df_contacto = pd.DataFrame(
    filas_contacto,
    columns=["id_universidad", "contacto", "ESA"]
)

df_observaciones = pd.DataFrame(
    filas_observaciones,
    columns=["id_convenio", "observaciones"]
)

df_universidad = dfconvenio[[
    "id_universidad","Universidad","INICIO","VIGENCIA","CIUDAD",
    "¿Católica?","¿SE USA?","Página Web",
    "ID_pais","ID_region","ID_giro","campus_id"
]]

# -----------------------------
# EXPORTAR
# -----------------------------

with pd.ExcelWriter(archivo_salida) as writer:
    df_universidad.to_excel(writer, sheet_name="universidad", index=False)
    df_convenio.to_excel(writer, sheet_name="convenio", index=False)

    df_tipo.drop(columns=["tipo_norm"]).to_excel(writer, sheet_name="tipo", index=False)
    df_escuela.drop(columns=["escuela_norm"]).to_excel(writer, sheet_name="escuela", index=False)

    df_conv_cve.to_excel(writer, sheet_name="convenio_cve", index=False)
    df_observaciones.to_excel(writer, sheet_name="observaciones", index=False)

    # 🔥 NUEVOS CATÁLOGOS
    df_pais.drop(columns=["pais_norm"]).to_excel(writer, sheet_name="pais", index=False)
    df_region.drop(columns=["region_norm"]).to_excel(writer, sheet_name="region", index=False)
    df_giro.drop(columns=["giro_norm"]).to_excel(writer, sheet_name="giro", index=False)
    df_campus.drop(columns=["campus_norm"]).to_excel(writer, sheet_name="campus", index=False)
    df_contacto.to_excel(writer, sheet_name="contacto", index=False)